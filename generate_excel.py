import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Carga Masiva OT"

# Setup styles
header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
align_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Headers on Row 6
headers = [
    "Número de OT",
    "RUC del Cliente",
    "Razón Social del Cliente",
    "Admin. de Contrato",
    "Partida",
    "Origen",
    "Destino / Dirección",
    "Fecha Requerida (YYYY-MM-DD)",
    "ID o SKU",
    "Descripción",
    "Cantidad",
    "Peso Unitario",
    "Peso Total"
]

# Insert Headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center
    cell.border = thin_border
    
    # Adjust column width
    ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 5, 20)

# Add some sample data on Row 7 and 8
sample_data = [
    ["OT-2023-0010", "10448588381", "JRM Empresa", "Demo Demo1", "1234", "Planta Chilca", "Lima, Sur", "2023-12-31", "SKU-001", "Cajas de Cartón", 100, 2.5, 250.0],
    ["OT-2023-0010", "10448588381", "JRM Empresa", "Demo Demo1", "1234", "Planta Chilca", "Lima, Sur", "2023-12-31", "SKU-002", "Pallets de Madera", 10, 20.0, 200.0],
]

for row_idx, row_data in enumerate(sample_data, 7):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

# Add Logo (Make sure we have PIL installed and the image available)
try:
    img = Image("public/logo-jrm.png")
    # Resize image to fit nicely
    img.width = 150
    img.height = 45
    ws.add_image(img, "A1")
except Exception as e:
    print("Could not add image:", e)

# Add title
title_cell = ws.cell(row=2, column=3, value="PLANTILLA DE CARGA MASIVA - ÓRDENES DE TRABAJO (OT)")
title_cell.font = Font(size=16, bold=True, color="002855")
ws.merge_cells("C2:G3")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Instructions
instr_cell = ws.cell(row=4, column=1, value="Instrucciones: Rellene los datos a partir de la fila 7. Para múltiples ítems, repita el 'Número de OT' en filas contiguas con distintos SKU.")
instr_cell.font = Font(italic=True, color="666666")
ws.merge_cells("A4:H4")

wb.save("public/Plantilla_Carga_Masiva_OT.xlsx")
print("Template saved to public/Plantilla_Carga_Masiva_OT.xlsx")
